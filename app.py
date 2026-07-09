# LESCO Bill Checker Portal - Rebuild Trigger
from flask import Flask, request, jsonify, render_template, send_file
import os
import openpyxl
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
import uuid
import logging
from scraper import get_bill_details

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = str(uuid.uuid4())

# In-memory storage for processed batches (in production, use Redis/db)
BATCH_CACHE = {}

# Clean temp folders (use cross-platform writeable system temp)
import tempfile
TEMP_DIR = os.path.join(tempfile.gettempdir(), 'temp')
os.makedirs(TEMP_DIR, exist_ok=True)

def find_reference_column(sheet):
    """
    Scans the header row (row 1) to find columns related to Reference Numbers or Customer IDs.
    """
    for col_idx in range(1, sheet.max_column + 1):
        cell_val = str(sheet.cell(row=1, column=col_idx).value or '').strip().lower()
        if any(keyword in cell_val for keyword in ['reference', 'ref', 'ref_no', 'ref no', 'refno', 'customer id', 'customer_id', 'consumer id', 'consumer_id', 'bill id', 'account']):
            logger.info(f"Detected Reference column at index {col_idx} with header '{sheet.cell(row=1, column=col_idx).value}'")
            return col_idx
    # Default to first column if no matches found
    logger.warning("No explicit reference header found. Defaulting to Column 1.")
    return 1

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/check_single', methods=['POST'])
def check_single():
    data = request.get_json() or {}
    ref_no = data.get('ref_no')
    if not ref_no:
        return jsonify({"status": "Error", "message": "Reference number is required."}), 400
        
    result = get_bill_details(ref_no)
    return jsonify(result)

@app.route('/view_bill/<ref_no>')
def view_bill_rendered(ref_no):
    clean_ref = ref_no.replace(" ", "").replace("-", "")
    import tempfile
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, f"{clean_ref}.html")
    
    # Try reading from system temp first (if cached)
    if os.path.exists(file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                html_content = f.read()
            return html_content, 200, {"Content-Type": "text/html"}
        except Exception as e:
            logger.warning(f"Failed to read cached temp bill: {e}")
            
    # Fallback: Dynamically fetch html from PITC
    from scraper import fetch_bill_html
    html, err = fetch_bill_html(clean_ref)
    if not err:
        # Inject base tag for PITC assets
        head_idx = html.lower().find("<head>")
        if head_idx != -1:
            insert_pos = head_idx + 6
            html = html[:insert_pos] + '\n    <base href="https://bill.pitc.com.pk/">' + html[insert_pos:]
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(html)
        except Exception as e:
            logger.warning(f"Failed to write bill to temp: {e}")
        return html, 200, {"Content-Type": "text/html"}
        
    return f"Bill details could not be retrieved: {err}", 404


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"status": "Error", "message": "No file uploaded."}), 400
        
    uploaded_file = request.files['file']
    if uploaded_file.filename == '':
        return jsonify({"status": "Error", "message": "Empty filename."}), 400
        
    if not uploaded_file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"status": "Error", "message": "Invalid file format. Please upload an Excel sheet (.xlsx, .xls)."}), 400
        
    try:
        # Read Excel in-memory
        file_bytes = uploaded_file.read()
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        
        # Locate the column with reference numbers
        ref_col = find_reference_column(sheet)
        
        # Read reference numbers
        inputs = []
        row_map = {}  # Keep track of Excel row index for each input
        
        for row_idx in range(2, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=ref_col).value
            if val is not None:
                val_str = str(val).strip()
                if val_str:
                    inputs.append(val_str)
                    row_map[val_str] = row_idx
                    
        if not inputs:
            return jsonify({"status": "Error", "message": "No data rows found in the Excel sheet."}), 400
            
        # Limit to maximum 30 reference numbers per batch to prevent serverless function timeout
        if len(inputs) > 30:
            return jsonify({"status": "Error", "message": "Batch limit exceeded. You can process up to 30 reference numbers at a time on the free tier."}), 400
            
        logger.info(f"Loaded {len(inputs)} reference numbers for batch processing.")
        
        # Max workers set to 1 (sequential) to prevent proxy concurrency overload and rate limits
        results = []
        max_workers = 1
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_input = {executor.submit(get_bill_details, item): item for item in inputs}
            for future in as_completed(future_to_input):
                orig_input = future_to_input[future]
                try:
                    result = future.result()
                    results.append(result)
                except Exception as exc:
                    logger.error(f"Execution error for {orig_input}: {exc}")
                    results.append({
                        "input_reference": orig_input,
                        "status": "Error",
                        "message": f"Processing exception: {str(exc)}"
                    })
                    
        # Cache results in memory mapped to a unique batch ID
        batch_id = str(uuid.uuid4())
        BATCH_CACHE[batch_id] = {
            "results": results,
            "original_file": file_bytes,
            "ref_column_index": ref_col
        }
        
        return jsonify({
            "status": "Success",
            "batch_id": batch_id,
            "total_items": len(inputs),
            "processed_items": len(results),
            "results": results
        })
        
    except Exception as e:
        logger.error(f"Error processing upload: {e}")
        return jsonify({"status": "Error", "message": f"Failed to process Excel file: {str(e)}"}), 500

@app.route('/api/download/<batch_id>', methods=['GET'])
def download_results(batch_id):
    batch_data = BATCH_CACHE.get(batch_id)
    if not batch_data:
        return "Batch not found or expired.", 404
        
    try:
        # Load the original file
        wb = openpyxl.load_workbook(BytesIO(batch_data["original_file"]))
        sheet = wb.active
        
        # Check current columns and prepare headers to append
        # We append headers to the right of existing columns
        start_col = sheet.max_column + 2  # Leave 1 column empty for visual separation
        
        headers_to_add = [
            "Fetched Reference", "Consumer ID", "Consumer Name", "Bill Month",
            "Due Date", "Payable Within Due Date", "Payable After Due Date",
            "Payment Status", "Amount Paid", "Payment Date", "Scraping Status", "Bill View Link"
        ]
        
        # Write extra headers
        sheet.cell(row=1, column=start_col - 1, value="|").font = openpyxl.styles.Font(bold=True)
        for idx, h in enumerate(headers_to_add):
            cell = sheet.cell(row=1, column=start_col + idx, value=h)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
            
        # Map inputs to outputs
        results_map = {res["input_reference"]: res for res in batch_data["results"]}
        ref_col = batch_data["ref_column_index"]
        
        # Fill in the data row by row
        for row_idx in range(2, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=ref_col).value
            sheet.cell(row=row_idx, column=start_col - 1, value="|") # separator
            
            if val is not None:
                val_str = str(val).strip()
                res = results_map.get(val_str)
                if res:
                    if res.get("message") == "Success":
                        sheet.cell(row=row_idx, column=start_col, value=res.get("reference_no", ""))
                        sheet.cell(row=row_idx, column=start_col + 1, value=res.get("consumer_id", ""))
                        sheet.cell(row=row_idx, column=start_col + 2, value=res.get("consumer_name", ""))
                        sheet.cell(row=row_idx, column=start_col + 3, value=res.get("bill_month", ""))
                        sheet.cell(row=row_idx, column=start_col + 4, value=res.get("due_date", ""))
                        sheet.cell(row=row_idx, column=start_col + 5, value=res.get("payable_within_due_date", 0))
                        sheet.cell(row=row_idx, column=start_col + 6, value=res.get("payable_after_due_date", 0))
                        sheet.cell(row=row_idx, column=start_col + 7, value=res.get("payment_status", ""))
                        sheet.cell(row=row_idx, column=start_col + 8, value=res.get("amount_paid", 0))
                        sheet.cell(row=row_idx, column=start_col + 9, value=res.get("payment_date", ""))
                        sheet.cell(row=row_idx, column=start_col + 10, value="Success")
                        
                        bill_link = res.get("bill_link")
                        if bill_link:
                            abs_url = f"{request.host_url.rstrip('/')}{bill_link}"
                            cell = sheet.cell(row=row_idx, column=start_col + 11, value="View Bill")
                            cell.hyperlink = abs_url
                            cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
                    else:
                        sheet.cell(row=row_idx, column=start_col + 10, value=res.get("message", "Failed"))
                else:
                    sheet.cell(row=row_idx, column=start_col + 10, value="Not Processed")
                    
        # Write to bytes stream
        out_stream = BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        
        return send_file(
            out_stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="LESCO_Bill_Checker_Results.xlsx"
        )
        
    except Exception as e:
        logger.error(f"Error compiling Excel download: {e}")
        return f"Failed to compile Excel download: {str(e)}", 500

if __name__ == '__main__':
    # Default to port 7860 for Hugging Face Spaces compatibility
    port = int(os.environ.get("PORT", 7860))
    app.run(host='0.0.0.0', port=port, debug=False)
